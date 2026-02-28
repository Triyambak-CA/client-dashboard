import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import copy

# ── Colour palette ──────────────────────────────────────────────
HDR_BG   = "1F3864"   # dark navy  – header row background
HDR_FG   = "FFFFFF"   # white      – header row text
SEC_BG   = "D9E1F2"   # light blue – section divider background
SEC_FG   = "1F3864"   # navy       – section divider text
REQ_FG   = "C00000"   # red        – required field name
OPT_FG   = "375623"   # dark green – optional field name
IND_BG   = "FFF2CC"   # pale amber – individual-only rows background
IND_LBL  = "7F6000"   # amber text – individual-only label
ALT_BG   = "F2F2F2"   # light grey – alternating rows

def hdr_style():
    return Font(bold=True, color=HDR_FG, size=10), PatternFill("solid", fgColor=HDR_BG)

def sec_style():
    return Font(bold=True, color=SEC_FG, size=10), PatternFill("solid", fgColor=SEC_BG)

def req_font():
    return Font(bold=True, color=REQ_FG, size=10)

def opt_font():
    return Font(color=OPT_FG, size=10)

def ind_fill():
    return PatternFill("solid", fgColor=IND_BG)

def alt_fill():
    return PatternFill("solid", fgColor=ALT_BG)

def plain_font():
    return Font(size=10)

def write_row(ws, row_num, values, fonts=None, fills=None, bold_first=False):
    for col, val in enumerate(values, 1):
        cell = ws.cell(row=row_num, column=col, value=val)
        cell.alignment = Alignment(wrap_text=True, vertical="top")
        if fonts and col - 1 < len(fonts) and fonts[col - 1]:
            cell.font = fonts[col - 1]
        if fills and col - 1 < len(fills) and fills[col - 1]:
            cell.fill = fills[col - 1]

def set_col_widths(ws, widths):
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

def freeze_and_filter(ws):
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions

# ── Build workbook ──────────────────────────────────────────────
wb = openpyxl.Workbook()
wb.remove(wb.active)   # remove default sheet

COLS = ["Field Name", "Data Type", "Required?", "Description / Notes", "Example Value"]

# ═══════════════════════════════════════════════════════════════
# README
# ═══════════════════════════════════════════════════════════════
readme = wb.create_sheet("README")
readme.column_dimensions["A"].width = 90
readme_lines = [
    ("HOW TO USE THIS FILE", True),
    ("", False),
    ("This file contains the proposed database schema for your CA firm's client management system.", False),
    ("Review each sheet, add/remove/rename fields, then share back.", False),
    ("", False),
    ("ARCHITECTURE — KEY DESIGN DECISION", True),
    ("", False),
    ("All entities (Individuals, Companies, Firms, LLPs, HUFs, Trusts, AOP/BOI) live in Sheet 1 — Clients (Master).", False),
    ("There is NO separate Persons/KYC sheet. Individual clients carry their own KYC fields (Aadhaar, DIN, DSC, etc.)", False),
    ("in the same record, visible only when Constitution = Individual.", False),
    ("", False),
    ("This means one person who is:", False),
    ("  • A direct client (personal IT returns)", False),
    ("  • A Director in Company A", False),
    ("  • A Partner in Firm B", False),
    ("  • A Shareholder in Company C", False),
    ("has exactly ONE Client ID. All roles (Director, Partner, Shareholder) link back to that single Client ID.", False),
    ("No duplication. Change their Aadhaar once — it reflects everywhere.", False),
    ("", False),
    ("SHEET GUIDE", True),
    ("", False),
    ("Sheet 1 — Clients (Master)       : One row per entity. Individuals carry KYC fields. Constitution drives which fields apply.", False),
    ("Sheet 2 — GST Registrations      : One row per GSTIN. Multiple rows per client for multiple state registrations.", False),
    ("Sheet 3 — Directors              : Links Individual Client IDs → Company Client IDs with designation & dates.", False),
    ("Sheet 4 — Shareholders           : Share ownership records. Individual or entity shareholders.", False),
    ("Sheet 5 — Partners (Firm/LLP)    : Links Individual Client IDs → Firm/LLP Client IDs with profit sharing ratios.", False),
    ("Sheet 6 — Bank Accounts          : Multiple bank accounts per client.", False),
    ("Sheet 7 — EPF/ESI Registrations  : EPF/ESI credentials per establishment.", False),
    ("Sheet 8 — Other Registrations    : TAN, MSME, IEC, FSSAI, Professional Tax, etc.", False),
    ("", False),
    ("COLOUR CODING IN EACH SHEET", True),
    ("", False),
    ("Red field name    = Required (must fill)", False),
    ("Green field name  = Optional (fill if available)", False),
    ("Amber row         = Individual-only field (skip for Companies/Firms/LLPs)", False),
    ("", False),
    ("WHAT TO DO", True),
    ("", False),
    ("1. Add fields    — Add a new row with field name, type, and description.", False),
    ("2. Remove fields — Delete rows you don't need.", False),
    ("3. Rename fields — Change the Field Name to your preferred terminology.", False),
    ("4. Change types  — Update Data Type if needed.", False),
    ("5. Share back    — Send the modified file and we'll build the database accordingly.", False),
]
for i, (text, bold) in enumerate(readme_lines, 1):
    cell = readme.cell(row=i, column=1, value=text)
    cell.font = Font(bold=bold, size=10 if not bold else 11)
    cell.alignment = Alignment(wrap_text=True)


# ═══════════════════════════════════════════════════════════════
# SHEET 1 — Clients (Master)
# ═══════════════════════════════════════════════════════════════
ws1 = wb.create_sheet("1 - Clients (Master)")
set_col_widths(ws1, [32, 18, 10, 52, 28])

# Header
r = 1
write_row(ws1, r, COLS,
          fonts=[Font(bold=True, color=HDR_FG, size=10)] * 5,
          fills=[PatternFill("solid", fgColor=HDR_BG)] * 5)

def sec(ws, row, label):
    write_row(ws, row, [label, None, None, None, None],
              fonts=[Font(bold=True, color=SEC_FG, size=10)] * 5,
              fills=[PatternFill("solid", fgColor=SEC_BG)] * 5)
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=5)

def ind_sec(ws, row, label):
    write_row(ws, row, [label, None, None, None, None],
              fonts=[Font(bold=True, color=IND_LBL, size=10)] * 5,
              fills=[PatternFill("solid", fgColor=IND_BG)] * 5)
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=5)

def field(ws, row, name, dtype, req, desc, example, individual_only=False):
    name_font = req_font() if req == "Yes" else opt_font()
    if individual_only:
        name_font = Font(bold=(req == "Yes"), color=(REQ_FG if req == "Yes" else IND_LBL), size=10)
    fill = ind_fill() if individual_only else None
    fills = [fill] * 5 if individual_only else None
    write_row(ws, row, [name, dtype, req, desc, example],
              fonts=[name_font, plain_font(), plain_font(), plain_font(), plain_font()],
              fills=fills)

r += 1; sec(ws1, r, "── IDENTITY")
r += 1; field(ws1, r, "Client ID",           "Auto (UUID)",     "Yes", "System generated unique ID — used everywhere as the primary key", "auto-generated")
r += 1; field(ws1, r, "PAN",                  "Text (10 char)",  "Yes", "Permanent Account Number — primary business identifier", "AASCM6627L")
r += 1; field(ws1, r, "Constitution",         "Dropdown",        "Yes", "Individual | Partnership Firm | LLP | Company | HUF | Trust | AOP | BOI", "Company")
r += 1; field(ws1, r, "Display Name",         "Text",            "Yes", "Short name used in the dashboard", "Mierae Solar Energy")
r += 1; field(ws1, r, "Legal Name",           "Text",            "Yes", "Name as per PAN / MCA / Partnership Deed records", "Mierae Solar Energy Private Limited")
r += 1; field(ws1, r, "Date of Incorporation / Birth", "Date",   "Yes", "DOI for Companies/LLPs/Firms, DOB for Individuals", "15/03/1990")
r += 1; field(ws1, r, "CIN / LLPIN",          "Text (21 char)",  "No",  "Company Identification Number (Companies) or LLP Identification Number (LLPs)", "U40100DL2015PTC285931")
r += 1; field(ws1, r, "TAN",                  "Text (10 char)",  "No",  "Tax Deduction Account Number — for entities with TDS obligations", "DELD01234E")
r += 1; field(ws1, r, "Is Direct Client?",    "Yes/No",          "Yes", "Yes = active CA client. No = registered only for KYC/compliance linkage (e.g. director of a client company)", "Yes")
r += 1; field(ws1, r, "Is Active?",           "Yes/No",          "Yes", "Mark inactive/former clients without deleting", "Yes")
r += 1; field(ws1, r, "Is on Retainer?",      "Yes/No",          "Yes", "Regular retainer client — for filtering and billing", "Yes")
r += 1; field(ws1, r, "Client Since",         "Date",            "No",  "When did this client join the firm", "01/04/2020")

r += 1; sec(ws1, r, "── INDIVIDUAL KYC  [Fields below apply only when Constitution = Individual]")
r += 1; ind_sec(ws1, r, "   ↳ Personal Identity")
r += 1; field(ws1, r, "Father's Name",        "Text",            "No",  "As per PAN / Aadhaar records", "Suresh Kumar Gupta",   individual_only=True)
r += 1; field(ws1, r, "Mother's Name",        "Text",            "No",  "Optional — required for some filings", None,                  individual_only=True)
r += 1; field(ws1, r, "Gender",               "Dropdown",        "No",  "Male | Female | Other", "Male",                                individual_only=True)
r += 1; field(ws1, r, "Nationality",          "Text",            "No",  "Default: Indian. Specify for NRIs / foreign nationals", "Indian", individual_only=True)
r += 1; ind_sec(ws1, r, "   ↳ KYC Numbers")
r += 1; field(ws1, r, "Aadhaar No.",          "Text (12 char)",  "No",  "Store masked — last 4 digits visible (XXXX-XXXX-1234)", "XXXX-XXXX-5678",      individual_only=True)
r += 1; field(ws1, r, "DIN",                  "Text (8 char)",   "No",  "Director Identification Number — fill if this individual is / was a director", "01234567",            individual_only=True)
r += 1; field(ws1, r, "Passport No.",         "Text",            "No",  "For NRIs and foreign nationals", None,                  individual_only=True)
r += 1; field(ws1, r, "Passport Expiry",      "Date",            "No",  "Alert before expiry", None,                            individual_only=True)
r += 1; ind_sec(ws1, r, "   ↳ MCA v3 (Personal Login)")
r += 1; field(ws1, r, "MCA User ID",          "Text",            "No",  "Personal MCA v3 portal login", None,                  individual_only=True)
r += 1; field(ws1, r, "MCA Password",         "Password",        "No",  "Personal MCA v3 portal password", None,               individual_only=True)
r += 1; ind_sec(ws1, r, "   ↳ DSC (Digital Signature Certificate)")
r += 1; field(ws1, r, "DSC Provider",         "Text",            "No",  "e.g. eMudhra, Sify, NSDL, Capricorn", "eMudhra",      individual_only=True)
r += 1; field(ws1, r, "DSC Expiry Date",      "Date",            "No",  "Alert before expiry", "31/12/2026",                   individual_only=True)
r += 1; field(ws1, r, "DSC Token Password",   "Password",        "No",  "USB token PIN", None,                                 individual_only=True)

r += 1; sec(ws1, r, "── CONTACT")
r += 1; field(ws1, r, "Primary Phone",        "Text",            "No",  "Main contact number", "9876543210")
r += 1; field(ws1, r, "Secondary Phone",      "Text",            "No",  "Alternate contact number", "9876543211")
r += 1; field(ws1, r, "Primary Email",        "Email",           "No",  "Main email ID", "info@mieraesolar.com")
r += 1; field(ws1, r, "Secondary Email",      "Email",           "No",  "Alternate email", None)

r += 1; sec(ws1, r, "── ADDRESS")
r += 1; field(ws1, r, "Address Line 1",       "Text",            "No",  "Street / building / flat details", "123 Business Park")
r += 1; field(ws1, r, "Address Line 2",       "Text",            "No",  "Area / locality / landmark", "Connaught Place")
r += 1; field(ws1, r, "City",                 "Text",            "No",  None, "New Delhi")
r += 1; field(ws1, r, "State",                "Dropdown",        "No",  "All 28 states + 8 UTs", "Delhi")
r += 1; field(ws1, r, "Pin Code",             "Text (6 char)",   "No",  None, "110001")

r += 1; sec(ws1, r, "── IT / INCOME TAX PORTAL")
r += 1; field(ws1, r, "IT Portal User ID",                "Text",      "No", "Usually same as PAN", "AASCM6627L")
r += 1; field(ws1, r, "IT Portal Password",               "Password",  "No", "Login password for incometax.gov.in", "Naveen1234$")
r += 1; field(ws1, r, "IT Portal User ID for TDS Filing", "Text",      "No", "Usually same as TAN — for entities with TDS", "DELD01234E")
r += 1; field(ws1, r, "IT Password for TDS Filing",       "Password",  "No", "Sometimes different from main IT password", "Naveen1234$")
r += 1; field(ws1, r, "Password for 26AS",                "Password",  "No", "Usually DOB in ddmmyyyy format", "10111990")
r += 1; field(ws1, r, "Password for AIS/TIS",             "Password",  "No", "Usually PAN in lowercase + DOB", "aascm6627l10111990")

r += 1; sec(ws1, r, "── TRACES")
r += 1; field(ws1, r, "TRACES User ID (Deductor)",   "Text",     "No", "For TDS filing — deductor login", None)
r += 1; field(ws1, r, "TRACES Password (Deductor)",  "Password", "No", None, None)
r += 1; field(ws1, r, "TRACES User ID (Tax Payer)",  "Text",     "No", "For taxpayer login on TRACES", None)
r += 1; field(ws1, r, "TRACES Password (Tax Payer)", "Password", "No", None, None)

r += 1; sec(ws1, r, "── NOTES")
r += 1; field(ws1, r, "Internal Notes", "Long Text", "No", "Any remarks or special instructions for this client", None)

freeze_and_filter(ws1)


# ═══════════════════════════════════════════════════════════════
# SHEET 2 — GST Registrations  (was Sheet 3)
# ═══════════════════════════════════════════════════════════════
ws2 = wb.create_sheet("2 - GST Registrations")
set_col_widths(ws2, [32, 18, 10, 52, 28])
r = 1
write_row(ws2, r, COLS,
          fonts=[Font(bold=True, color=HDR_FG, size=10)] * 5,
          fills=[PatternFill("solid", fgColor=HDR_BG)] * 5)
r += 1; sec(ws2, r, "── REGISTRATION DETAILS")
r += 1; field(ws2, r, "GST Reg. ID",        "Auto (UUID)",    "Yes", "System generated", "auto-generated")
r += 1; field(ws2, r, "Client ID",           "Link → Clients", "Yes", "Links to the parent client (Sheet 1)", "auto-linked")
r += 1; field(ws2, r, "GSTIN",               "Text (15 char)", "Yes", "GST Identification Number — unique per registration", "07AASCM6627L1ZQ")
r += 1; field(ws2, r, "State",               "Dropdown",       "Yes", "Auto-derived from GSTIN first 2 digits", "Delhi")
r += 1; field(ws2, r, "State Code",          "Text (2 char)",  "Yes", "01–37 — auto filled from GSTIN", "07")
r += 1; field(ws2, r, "Registration Type",   "Dropdown",       "No",  "Regular | Composition | QRMP | SEZ Unit | SEZ Developer | Casual | Non-Resident", "Regular")
r += 1; field(ws2, r, "Registration Date",   "Date",           "No",  "Date of GST registration", "01/07/2017")
r += 1; field(ws2, r, "Cancellation Date",   "Date",           "No",  "Fill only if registration is cancelled", None)
r += 1; field(ws2, r, "Is Active?",          "Yes/No",         "Yes", None, "Yes")
r += 1; sec(ws2, r, "── GST PORTAL CREDENTIALS")
r += 1; field(ws2, r, "GST User ID",         "Text",           "No",  "Login ID for GST portal", "MIERAESOLAR2025")
r += 1; field(ws2, r, "GST Password",        "Password",       "No",  None, "Gst@2025")
r += 1; sec(ws2, r, "── E-WAY BILL")
r += 1; field(ws2, r, "EWB User ID",         "Text",           "No",  "E-Way Bill portal login", None)
r += 1; field(ws2, r, "EWB Password",        "Password",       "No",  None, None)
r += 1; field(ws2, r, "EWB API User ID",     "Text",           "No",  "For API-based E-Way Bill generation", "DSNHP3424014000")
r += 1; field(ws2, r, "EWB API Password",    "Password",       "No",  None, "Admin@2025")
r += 1; sec(ws2, r, "── AUTHORISED SIGNATORY")
r += 1; field(ws2, r, "Authorised Signatory", "Text",          "No",  "Name of the authorised signatory for this GSTIN", "Rajesh Gupta")
r += 1; field(ws2, r, "Signatory PAN",        "Text",          "No",  "PAN of the authorised signatory", "ABCPG1234D")
r += 1; sec(ws2, r, "── NOTES")
r += 1; field(ws2, r, "Notes", "Long Text", "No", "e.g. Filing frequency, special instructions", None)
freeze_and_filter(ws2)


# ═══════════════════════════════════════════════════════════════
# SHEET 3 — Directors  (was Sheet 4)
# ═══════════════════════════════════════════════════════════════
ws3 = wb.create_sheet("3 - Directors")
set_col_widths(ws3, [32, 18, 10, 52, 28])
r = 1
write_row(ws3, r, COLS,
          fonts=[Font(bold=True, color=HDR_FG, size=10)] * 5,
          fills=[PatternFill("solid", fgColor=HDR_BG)] * 5)
r += 1; field(ws3, r, "Director Record ID",     "Auto (UUID)",    "Yes", "System generated", "auto-generated")
r += 1; field(ws3, r, "Company Client ID",       "Link → Clients", "Yes", "Client ID of the Company (Constitution = Company)", "auto-linked")
r += 1; field(ws3, r, "Individual Client ID",    "Link → Clients", "Yes", "Client ID of the Director (Constitution = Individual) — single source of truth for this person's KYC", "auto-linked")
r += 1; field(ws3, r, "DIN",                     "Text (8 char)",  "Yes", "Director Identification Number — auto-fetched from the Individual's Client record", "01234567")
r += 1; field(ws3, r, "Designation",             "Dropdown",       "Yes", "Director | Managing Director | Whole-time Director | Independent Director | Nominee Director | Additional Director", "Director")
r += 1; field(ws3, r, "Date of Appointment",     "Date",           "No",  None, "01/04/2015")
r += 1; field(ws3, r, "Date of Cessation",       "Date",           "No",  "Leave blank if currently active", None)
r += 1; field(ws3, r, "Is Active?",              "Yes/No",         "Yes", None, "Yes")
r += 1; field(ws3, r, "Is KMP?",                 "Yes/No",         "No",  "Key Managerial Personnel", "Yes")
r += 1; field(ws3, r, "Notes",                   "Long Text",      "No",  None, None)
freeze_and_filter(ws3)


# ═══════════════════════════════════════════════════════════════
# SHEET 4 — Shareholders  (was Sheet 5)
# ═══════════════════════════════════════════════════════════════
ws4 = wb.create_sheet("4 - Shareholders")
set_col_widths(ws4, [32, 18, 10, 52, 28])
r = 1
write_row(ws4, r, COLS,
          fonts=[Font(bold=True, color=HDR_FG, size=10)] * 5,
          fills=[PatternFill("solid", fgColor=HDR_BG)] * 5)
r += 1; field(ws4, r, "Shareholder ID",         "Auto (UUID)",    "Yes", "System generated", "auto-generated")
r += 1; field(ws4, r, "Company Client ID",       "Link → Clients", "Yes", "Client ID of the Company whose shares are held", "auto-linked")
r += 1; field(ws4, r, "Holder Type",             "Dropdown",       "Yes", "Individual | Company | Trust | HUF | LLP", "Individual")
r += 1; field(ws4, r, "Individual Client ID",    "Link → Clients", "No",  "Fill if Holder Type = Individual — links to that person's single Client record", "auto-linked")
r += 1; field(ws4, r, "Holding Entity Client ID","Link → Clients", "No",  "Fill if Holder Type = Company / Trust / LLP / HUF — links to that entity's Client record", None)
r += 1; field(ws4, r, "Holder Name",             "Text",           "Yes", "Name as per share records — auto-fetched from linked Client record", "Rajesh Kumar Gupta")
r += 1; field(ws4, r, "PAN of Holder",           "Text",           "No",  "Auto-fetched from linked Client record", "ABCPG1234D")
r += 1; field(ws4, r, "Share Type",              "Dropdown",       "No",  "Equity | Preference | CCPS | OCPS", "Equity")
r += 1; field(ws4, r, "Number of Shares",        "Number",         "No",  None, "5000")
r += 1; field(ws4, r, "Face Value (₹)",          "Number",         "No",  "Per share face value", "10")
r += 1; field(ws4, r, "Percentage %",            "Decimal",        "No",  "Shareholding percentage", "50.00")
r += 1; field(ws4, r, "Date Acquired",           "Date",           "No",  None, "01/04/2015")
r += 1; field(ws4, r, "Is Active?",              "Yes/No",         "Yes", None, "Yes")
r += 1; field(ws4, r, "Notes",                   "Long Text",      "No",  None, None)
freeze_and_filter(ws4)


# ═══════════════════════════════════════════════════════════════
# SHEET 5 — Partners (Firm-LLP)  (was Sheet 6)
# ═══════════════════════════════════════════════════════════════
ws5 = wb.create_sheet("5 - Partners (Firm-LLP)")
set_col_widths(ws5, [32, 18, 10, 52, 28])
r = 1
write_row(ws5, r, COLS,
          fonts=[Font(bold=True, color=HDR_FG, size=10)] * 5,
          fills=[PatternFill("solid", fgColor=HDR_BG)] * 5)
r += 1; field(ws5, r, "Partner Record ID",      "Auto (UUID)",    "Yes", "System generated", "auto-generated")
r += 1; field(ws5, r, "Firm / LLP Client ID",   "Link → Clients", "Yes", "Client ID of the Firm or LLP (Constitution = Partnership Firm or LLP)", "auto-linked")
r += 1; field(ws5, r, "Individual Client ID",   "Link → Clients", "Yes", "Client ID of the Partner (Constitution = Individual) — single source of truth for this person's KYC", "auto-linked")
r += 1; field(ws5, r, "Role",                   "Dropdown",       "Yes", "Partner | Designated Partner | Managing Partner | Sleeping Partner | Minor Partner", "Designated Partner")
r += 1; field(ws5, r, "Profit Sharing Ratio %", "Decimal",        "No",  "As per partnership deed / LLP agreement", "50.00")
r += 1; field(ws5, r, "Capital Contribution (₹)","Number",        "No",  None, "500000")
r += 1; field(ws5, r, "Date of Joining",        "Date",           "No",  None, "01/04/2015")
r += 1; field(ws5, r, "Date of Exit",           "Date",           "No",  "Leave blank if currently active", None)
r += 1; field(ws5, r, "Is Active?",             "Yes/No",         "Yes", None, "Yes")
r += 1; field(ws5, r, "Notes",                  "Long Text",      "No",  None, None)
freeze_and_filter(ws5)


# ═══════════════════════════════════════════════════════════════
# SHEET 6 — Bank Accounts  (was Sheet 7)
# ═══════════════════════════════════════════════════════════════
ws6 = wb.create_sheet("6 - Bank Accounts")
set_col_widths(ws6, [32, 18, 10, 52, 28])
r = 1
write_row(ws6, r, COLS,
          fonts=[Font(bold=True, color=HDR_FG, size=10)] * 5,
          fills=[PatternFill("solid", fgColor=HDR_BG)] * 5)
r += 1; field(ws6, r, "Bank Account ID",      "Auto (UUID)",    "Yes", "System generated", "auto-generated")
r += 1; field(ws6, r, "Client ID",            "Link → Clients", "Yes", "The client this account belongs to", "auto-linked")
r += 1; field(ws6, r, "Bank Name",            "Text",           "Yes", None, "State Bank of India")
r += 1; field(ws6, r, "Account Number",       "Text",           "Yes", None, "20001791050000999")
r += 1; field(ws6, r, "IFSC Code",            "Text (11 char)", "Yes", "Bank branch IFSC", "SBIN0001234")
r += 1; field(ws6, r, "Branch Name",          "Text",           "No",  None, "Connaught Place")
r += 1; field(ws6, r, "Account Type",         "Dropdown",       "No",  "Current | Savings | Cash Credit | Overdraft | EEFC", "Current")
r += 1; field(ws6, r, "Is Primary?",          "Yes/No",         "Yes", "Mark the main operating account", "Yes")
r += 1; field(ws6, r, "Net Banking User ID",  "Text",           "No",  None, None)
r += 1; field(ws6, r, "Net Banking Password", "Password",       "No",  None, None)
r += 1; field(ws6, r, "Notes",                "Long Text",      "No",  None, None)
freeze_and_filter(ws6)


# ═══════════════════════════════════════════════════════════════
# SHEET 7 — EPF/ESI Registrations  (was Sheet 8)
# ═══════════════════════════════════════════════════════════════
ws7 = wb.create_sheet("7 - EPF_ESI Registrations")
set_col_widths(ws7, [32, 18, 10, 52, 28])
r = 1
write_row(ws7, r, COLS,
          fonts=[Font(bold=True, color=HDR_FG, size=10)] * 5,
          fills=[PatternFill("solid", fgColor=HDR_BG)] * 5)
r += 1; sec(ws7, r, "── REGISTRATION DETAILS")
r += 1; field(ws7, r, "EPF/ESI Reg. ID",              "Auto (UUID)",    "Yes", "System generated", "auto-generated")
r += 1; field(ws7, r, "Client ID",                     "Link → Clients", "Yes", "Links to the parent client", "auto-linked")
r += 1; field(ws7, r, "Registration Type",             "Dropdown",       "Yes", "EPF | ESI", "EPF")
r += 1; field(ws7, r, "State",                         "Dropdown",       "Yes", "State where this EPF/ESI establishment is registered", "Delhi")
r += 1; field(ws7, r, "Establishment Code / Reg. No.", "Text",           "Yes", "EPF Establishment Code or ESI Registration Number", "DLCPM0012345000")
r += 1; field(ws7, r, "Registration Date",             "Date",           "No",  "Date of EPF/ESI registration", "04/01/2020")
r += 1; field(ws7, r, "Cancellation Date",             "Date",           "No",  "Fill only if registration is cancelled", None)
r += 1; field(ws7, r, "Is Active?",                    "Yes/No",         "Yes", None, "Yes")
r += 1; sec(ws7, r, "── PORTAL CREDENTIALS")
r += 1; field(ws7, r, "Portal User ID",       "Text",     "No", "Login ID for EPFO / ESIC portal", "DLCPM0012345000")
r += 1; field(ws7, r, "Portal Password",      "Password", "No", None, "Epf@2025")
r += 1; field(ws7, r, "DSC Holder Name",      "Text",     "No", "Name of person whose DSC is used for filing", "Rajesh Gupta")
r += 1; field(ws7, r, "Authorised Signatory", "Text",     "No", "Name of the authorised signatory for this registration", "Rajesh Gupta")
r += 1; sec(ws7, r, "── NOTES")
r += 1; field(ws7, r, "Notes", "Long Text", "No", "e.g. Filing frequency, wage ceiling, special instructions", None)
freeze_and_filter(ws7)


# ═══════════════════════════════════════════════════════════════
# SHEET 8 — Other Registrations  (was Sheet 9)
# ═══════════════════════════════════════════════════════════════
ws8 = wb.create_sheet("8 - Other Registrations")
set_col_widths(ws8, [32, 18, 10, 52, 28])
r = 1
write_row(ws8, r, COLS,
          fonts=[Font(bold=True, color=HDR_FG, size=10)] * 5,
          fills=[PatternFill("solid", fgColor=HDR_BG)] * 5)
r += 1; field(ws8, r, "Registration ID",        "Auto (UUID)",    "Yes", "System generated", "auto-generated")
r += 1; field(ws8, r, "Client ID",              "Link → Clients", "Yes", "The client this registration belongs to", "auto-linked")
r += 1; field(ws8, r, "Registration Type",      "Dropdown",       "Yes", "TAN | MSME/Udyam | IEC | FSSAI | Professional Tax | Shops & Estab | Trade License | Drug License | Import Export Code | Others", "MSME/Udyam")
r += 1; field(ws8, r, "Registration Number",    "Text",           "Yes", "Unique registration/certificate number", "UDYAM-DL-01-0012345")
r += 1; field(ws8, r, "Registration Date",      "Date",           "No",  None, "15/06/2021")
r += 1; field(ws8, r, "Valid Until",            "Date",           "No",  "For licenses with expiry — system will alert before expiry", "14/06/2026")
r += 1; field(ws8, r, "Issuing Authority",      "Text",           "No",  "Government dept that issued this", "Ministry of MSME")
r += 1; field(ws8, r, "State / Jurisdiction",   "Dropdown",       "No",  "Relevant for state-specific registrations", "Delhi")
r += 1; field(ws8, r, "Portal User ID",         "Text",           "No",  "Login for this registration portal", None)
r += 1; field(ws8, r, "Portal Password",        "Password",       "No",  None, None)
r += 1; field(ws8, r, "Is Active?",             "Yes/No",         "Yes", None, "Yes")
r += 1; field(ws8, r, "Notes",                  "Long Text",      "No",  None, None)
freeze_and_filter(ws8)


# ── Save ────────────────────────────────────────────────────────
wb.save("CA_Client_Schema.xlsx")
print("Done — CA_Client_Schema.xlsx saved successfully.")
print(f"Sheets: {wb.sheetnames}")
