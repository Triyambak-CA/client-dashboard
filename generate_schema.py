import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

wb = openpyxl.Workbook()

# Color palette
HEADER_BG    = "1F3864"  # dark navy
HEADER_FG    = "FFFFFF"
SECTION_BG   = "D6E4F0"  # light blue
SECTION_FG   = "1F3864"
ALT_ROW      = "F2F7FB"
WHITE        = "FFFFFF"
REQUIRED_COL = "E74C3C"  # red for required
OPTIONAL_COL = "27AE60"  # green for optional

thin = Side(style='thin', color="CCCCCC")
border = Border(left=thin, right=thin, top=thin, bottom=thin)

def style_header(cell, text):
    cell.value = text
    cell.font = Font(bold=True, color=HEADER_FG, size=11)
    cell.fill = PatternFill("solid", fgColor=HEADER_BG)
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.border = border

def style_section(cell, text):
    cell.value = text
    cell.font = Font(bold=True, color=SECTION_FG, size=10)
    cell.fill = PatternFill("solid", fgColor=SECTION_BG)
    cell.alignment = Alignment(horizontal="left", vertical="center")
    cell.border = border

def style_row(cell, text, row_num, is_required=None):
    cell.value = text
    bg = ALT_ROW if row_num % 2 == 0 else WHITE
    cell.fill = PatternFill("solid", fgColor=bg)
    cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    cell.border = border
    if is_required is True:
        cell.font = Font(color=REQUIRED_COL, size=10)
    elif is_required is False:
        cell.font = Font(color=OPTIONAL_COL, size=10)
    else:
        cell.font = Font(size=10)

def create_sheet(wb, title, columns, data, col_widths):
    ws = wb.create_sheet(title=title)
    ws.row_dimensions[1].height = 30

    # Header row
    for col_idx, col_name in enumerate(columns, 1):
        style_header(ws.cell(row=1, column=col_idx), col_name)

    # Data rows
    row_num = 2
    for item in data:
        if item.get('_section'):
            # Section divider row - merge across all columns
            ws.merge_cells(start_row=row_num, start_column=1, end_row=row_num, end_column=len(columns))
            style_section(ws.cell(row=row_num, column=1), item['_section'])
            row_num += 1
            continue

        req = item.get('required')
        for col_idx, col_name in enumerate(columns, 1):
            key = col_name.lower().replace(' ', '_').replace('/', '_').replace('(', '').replace(')', '')
            # Try different key mappings
            val = item.get(col_name, item.get(key, ''))
            cell = ws.cell(row=row_num, column=col_idx)
            if col_name == 'Required?':
                style_row(cell, 'Yes' if req else 'No', row_num, req)
            else:
                style_row(cell, val, row_num)
        row_num += 1

    # Column widths
    for col_idx, width in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    ws.freeze_panes = "A2"
    return ws

COLS = ["Field Name", "Data Type", "Required?", "Description / Notes", "Example Value"]
WIDTHS = [28, 18, 12, 45, 30]

# ─────────────────────────────────────────────
# SHEET 1 — CLIENTS (MASTER)
# ─────────────────────────────────────────────
clients_data = [
    {'_section': '── IDENTITY'},
    {'Field Name': 'Client ID',         'Data Type': 'Auto (UUID)',    'required': True,  'Description / Notes': 'System generated unique ID',                         'Example Value': 'auto-generated'},
    {'Field Name': 'PAN',               'Data Type': 'Text (10 char)', 'required': True,  'Description / Notes': 'Permanent Account Number — primary identifier',       'Example Value': 'AASCM6627L'},
    {'Field Name': 'Display Name',      'Data Type': 'Text',           'required': True,  'Description / Notes': 'Short name used in the dashboard',                    'Example Value': 'Mierae Solar Energy'},
    {'Field Name': 'Legal Name',        'Data Type': 'Text',           'required': True,  'Description / Notes': 'Name as per PAN / MCA records',                       'Example Value': 'Mierae Solar Energy Private Limited'},
    {'Field Name': 'Constitution',      'Data Type': 'Dropdown',       'required': True,  'Description / Notes': 'Individual | Proprietorship | Partnership Firm | LLP | Private Ltd | Public Ltd | HUF | Trust | AOP | BOI', 'Example Value': 'Private Ltd'},
    {'Field Name': 'Date of Incorporation / Birth', 'Data Type': 'Date', 'required': False, 'Description / Notes': 'DOI for companies/LLP, DOB for individuals',        'Example Value': '15/03/1990'},
    {'Field Name': 'Financial Year End','Data Type': 'Dropdown',       'required': False, 'Description / Notes': '31-Mar | 30-Jun | 30-Sep | 31-Dec',                   'Example Value': '31-Mar'},
    {'_section': '── CONTACT'},
    {'Field Name': 'Primary Phone',     'Data Type': 'Text',           'required': False, 'Description / Notes': 'Main contact number',                                 'Example Value': '9876543210'},
    {'Field Name': 'Secondary Phone',   'Data Type': 'Text',           'required': False, 'Description / Notes': 'Alternate contact number',                            'Example Value': '9876543211'},
    {'Field Name': 'Primary Email',     'Data Type': 'Email',          'required': False, 'Description / Notes': 'Main email ID',                                       'Example Value': 'info@mieraesolar.com'},
    {'Field Name': 'Secondary Email',   'Data Type': 'Email',          'required': False, 'Description / Notes': 'Alternate email',                                     'Example Value': ''},
    {'_section': '── ADDRESS'},
    {'Field Name': 'Address Line 1',    'Data Type': 'Text',           'required': False, 'Description / Notes': 'Street / building details',                           'Example Value': '123 Business Park'},
    {'Field Name': 'Address Line 2',    'Data Type': 'Text',           'required': False, 'Description / Notes': 'Area / locality',                                     'Example Value': 'Connaught Place'},
    {'Field Name': 'City',              'Data Type': 'Text',           'required': False, 'Description / Notes': '',                                                     'Example Value': 'New Delhi'},
    {'Field Name': 'State',             'Data Type': 'Dropdown',       'required': False, 'Description / Notes': 'All 28 states + 8 UTs',                               'Example Value': 'Delhi'},
    {'Field Name': 'Pin Code',          'Data Type': 'Text (6 char)',  'required': False, 'Description / Notes': '',                                                     'Example Value': '110001'},
    {'_section': '── IT / INCOME TAX'},
    {'Field Name': 'IT Portal User ID', 'Data Type': 'Text',           'required': False, 'Description / Notes': 'Usually same as PAN',                                 'Example Value': 'AASCM6627L'},
    {'Field Name': 'IT Portal Password','Data Type': 'Password',       'required': False, 'Description / Notes': 'Login password for incometax.gov.in',                 'Example Value': 'Naveen1234$'},
    {'Field Name': 'IT Password for TDS Filing', 'Data Type': 'Password', 'required': False, 'Description / Notes': 'Sometimes different from main IT password',       'Example Value': ''},
    {'_section': '── TRACES'},
    {'Field Name': 'TRACES User ID (Deductor)',   'Data Type': 'Text',     'required': False, 'Description / Notes': 'For TDS filing — deductor login',                'Example Value': ''},
    {'Field Name': 'TRACES Password (Deductor)',  'Data Type': 'Password', 'required': False, 'Description / Notes': '',                                                'Example Value': ''},
    {'Field Name': 'TRACES User ID (Tax Payer)',  'Data Type': 'Text',     'required': False, 'Description / Notes': 'For taxpayer login on TRACES',                   'Example Value': ''},
    {'Field Name': 'TRACES Password (Tax Payer)', 'Data Type': 'Password', 'required': False, 'Description / Notes': '',                                                'Example Value': ''},
    {'_section': '── MCA / COMPANY (if applicable)'},
    {'Field Name': 'CIN',               'Data Type': 'Text (21 char)', 'required': False, 'Description / Notes': 'Company Identification Number (Companies/LLPs only)', 'Example Value': 'U40100DL2015PTC285931'},
    {'Field Name': 'MCA V3 User ID',    'Data Type': 'Text',           'required': False, 'Description / Notes': 'Login for MCA21 portal',                             'Example Value': ''},
    {'Field Name': 'MCA V3 Password',   'Data Type': 'Password',       'required': False, 'Description / Notes': '',                                                    'Example Value': ''},
    {'_section': '── OTHER PORTALS'},
    {'Field Name': 'Password for 26AS', 'Data Type': 'Password',       'required': False, 'Description / Notes': 'Usually PAN in lowercase + DOB or similar',          'Example Value': 'aascm6627l00011900'},
    {'Field Name': 'Password for AIS/TIS','Data Type': 'Password',     'required': False, 'Description / Notes': 'AIS/TIS portal password',                            'Example Value': ''},
    {'_section': '── TAN / TDS'},
    {'Field Name': 'TAN',               'Data Type': 'Text (10 char)', 'required': False, 'Description / Notes': 'Tax Deduction Account Number',                       'Example Value': '00011900'},
    {'Field Name': 'TAN Portal User ID','Data Type': 'Text',           'required': False, 'Description / Notes': '',                                                    'Example Value': ''},
    {'Field Name': 'TAN Portal Password','Data Type': 'Password',      'required': False, 'Description / Notes': '',                                                    'Example Value': ''},
    {'_section': '── EPF / ESI'},
    {'Field Name': 'EPF Establishment Code', 'Data Type': 'Text',      'required': False, 'Description / Notes': 'PF registration number',                             'Example Value': ''},
    {'Field Name': 'EPF Login',         'Data Type': 'Text',           'required': False, 'Description / Notes': '',                                                    'Example Value': ''},
    {'Field Name': 'EPF Password',      'Data Type': 'Password',       'required': False, 'Description / Notes': '',                                                    'Example Value': ''},
    {'Field Name': 'ESI Registration No.','Data Type': 'Text',         'required': False, 'Description / Notes': '17-digit ESI number',                                'Example Value': ''},
    {'Field Name': 'ESI Login',         'Data Type': 'Text',           'required': False, 'Description / Notes': '',                                                    'Example Value': ''},
    {'Field Name': 'ESI Password',      'Data Type': 'Password',       'required': False, 'Description / Notes': '',                                                    'Example Value': ''},
    {'_section': '── NOTES'},
    {'Field Name': 'Internal Notes',    'Data Type': 'Long Text',      'required': False, 'Description / Notes': 'Any remarks or special instructions for this client', 'Example Value': ''},
    {'Field Name': 'Client Since',      'Data Type': 'Date',           'required': False, 'Description / Notes': 'When did this client join the firm',                  'Example Value': '01/04/2020'},
    {'Field Name': 'Is Active?',        'Data Type': 'Yes/No',         'required': True,  'Description / Notes': 'To mark inactive/former clients without deleting',    'Example Value': 'Yes'},
]

# ─────────────────────────────────────────────
# SHEET 2 — PERSONS (KYC)
# ─────────────────────────────────────────────
persons_data = [
    {'_section': '── IDENTITY'},
    {'Field Name': 'Person ID',         'Data Type': 'Auto (UUID)',    'required': True,  'Description / Notes': 'System generated',                                    'Example Value': 'auto-generated'},
    {'Field Name': 'Full Name',         'Data Type': 'Text',           'required': True,  'Description / Notes': 'As per PAN card',                                     'Example Value': 'Rajesh Kumar Gupta'},
    {'Field Name': "Father's Name",     'Data Type': 'Text',           'required': False, 'Description / Notes': '',                                                    'Example Value': 'Suresh Kumar Gupta'},
    {'Field Name': "Mother's Name",     'Data Type': 'Text',           'required': False, 'Description / Notes': '',                                                    'Example Value': ''},
    {'Field Name': 'Date of Birth',     'Data Type': 'Date',           'required': False, 'Description / Notes': '',                                                    'Example Value': '15/03/1975'},
    {'Field Name': 'Gender',            'Data Type': 'Dropdown',       'required': False, 'Description / Notes': 'Male | Female | Other',                              'Example Value': 'Male'},
    {'Field Name': 'Nationality',       'Data Type': 'Text',           'required': False, 'Description / Notes': 'Default: Indian',                                     'Example Value': 'Indian'},
    {'_section': '── KYC NUMBERS'},
    {'Field Name': 'PAN (Individual)',  'Data Type': 'Text (10 char)', 'required': True,  'Description / Notes': 'Personal PAN — unique identifier for this person',    'Example Value': 'ABCPG1234D'},
    {'Field Name': 'Aadhaar No.',       'Data Type': 'Text (12 char)', 'required': False, 'Description / Notes': 'Store masked (last 4 digits visible)',                'Example Value': 'XXXX-XXXX-5678'},
    {'Field Name': 'DIN',               'Data Type': 'Text (8 char)',  'required': False, 'Description / Notes': 'Director Identification Number (if director)',         'Example Value': '01234567'},
    {'Field Name': 'Passport No.',      'Data Type': 'Text',           'required': False, 'Description / Notes': 'For foreign nationals or NRIs',                       'Example Value': ''},
    {'Field Name': 'Passport Expiry',   'Data Type': 'Date',           'required': False, 'Description / Notes': '',                                                    'Example Value': ''},
    {'_section': '── CONTACT'},
    {'Field Name': 'Phone',             'Data Type': 'Text',           'required': False, 'Description / Notes': '',                                                    'Example Value': '9876543210'},
    {'Field Name': 'Email',             'Data Type': 'Email',          'required': False, 'Description / Notes': '',                                                    'Example Value': 'rajesh@email.com'},
    {'_section': '── ADDRESS'},
    {'Field Name': 'Residential Address','Data Type': 'Text',          'required': False, 'Description / Notes': 'Full address as per Aadhaar/PAN',                     'Example Value': '45 Green Park, New Delhi'},
    {'Field Name': 'City',              'Data Type': 'Text',           'required': False, 'Description / Notes': '',                                                    'Example Value': 'New Delhi'},
    {'Field Name': 'State',             'Data Type': 'Dropdown',       'required': False, 'Description / Notes': '',                                                    'Example Value': 'Delhi'},
    {'Field Name': 'Pin Code',          'Data Type': 'Text (6 char)',  'required': False, 'Description / Notes': '',                                                    'Example Value': '110016'},
    {'_section': '── IT CREDENTIALS (Personal)'},
    {'Field Name': 'IT Portal Password','Data Type': 'Password',       'required': False, 'Description / Notes': 'Personal IT portal login',                            'Example Value': ''},
    {'Field Name': 'TRACES Password',   'Data Type': 'Password',       'required': False, 'Description / Notes': 'Personal TRACES login',                               'Example Value': ''},
    {'_section': '── DSC'},
    {'Field Name': 'DSC Provider',      'Data Type': 'Text',           'required': False, 'Description / Notes': 'e.g. eMudhra, Sify, NSDL',                           'Example Value': 'eMudhra'},
    {'Field Name': 'DSC Expiry Date',   'Data Type': 'Date',           'required': False, 'Description / Notes': 'Alert before expiry',                                 'Example Value': '31/12/2026'},
    {'Field Name': 'DSC Token Password','Data Type': 'Password',       'required': False, 'Description / Notes': 'USB token PIN',                                       'Example Value': ''},
    {'_section': '── NOTES'},
    {'Field Name': 'Notes',             'Data Type': 'Long Text',      'required': False, 'Description / Notes': 'Any additional information about this person',        'Example Value': ''},
]

# ─────────────────────────────────────────────
# SHEET 3 — GST REGISTRATIONS
# ─────────────────────────────────────────────
gst_data = [
    {'_section': '── REGISTRATION DETAILS'},
    {'Field Name': 'GST Reg. ID',       'Data Type': 'Auto (UUID)',    'required': True,  'Description / Notes': 'System generated',                                    'Example Value': 'auto-generated'},
    {'Field Name': 'Client ID',         'Data Type': 'Link → Clients', 'required': True,  'Description / Notes': 'Links to the parent client',                         'Example Value': 'auto-linked'},
    {'Field Name': 'GSTIN',             'Data Type': 'Text (15 char)', 'required': True,  'Description / Notes': 'GST Identification Number — unique per registration', 'Example Value': '07AASCM6627L1ZQ'},
    {'Field Name': 'State',             'Data Type': 'Dropdown',       'required': True,  'Description / Notes': 'Auto-derived from GSTIN first 2 digits',              'Example Value': 'Delhi'},
    {'Field Name': 'State Code',        'Data Type': 'Text (2 char)',  'required': True,  'Description / Notes': '01–37 — auto filled from GSTIN',                     'Example Value': '07'},
    {'Field Name': 'Registration Type', 'Data Type': 'Dropdown',       'required': False, 'Description / Notes': 'Regular | Composition | QRMP | SEZ Unit | SEZ Developer | Casual | Non-Resident', 'Example Value': 'Regular'},
    {'Field Name': 'Registration Date', 'Data Type': 'Date',           'required': False, 'Description / Notes': 'Date of GST registration',                            'Example Value': '01/07/2017'},
    {'Field Name': 'Cancellation Date', 'Data Type': 'Date',           'required': False, 'Description / Notes': 'Fill only if registration is cancelled',              'Example Value': ''},
    {'Field Name': 'Is Active?',        'Data Type': 'Yes/No',         'required': True,  'Description / Notes': '',                                                    'Example Value': 'Yes'},
    {'_section': '── GST PORTAL CREDENTIALS'},
    {'Field Name': 'GST User ID',       'Data Type': 'Text',           'required': False, 'Description / Notes': 'Login ID for GST portal',                             'Example Value': 'MIERAESOLAR2025'},
    {'Field Name': 'GST Password',      'Data Type': 'Password',       'required': False, 'Description / Notes': '',                                                    'Example Value': 'Gst@2025'},
    {'_section': '── E-WAY BILL'},
    {'Field Name': 'EWB User ID',       'Data Type': 'Text',           'required': False, 'Description / Notes': 'E-Way Bill portal login',                            'Example Value': ''},
    {'Field Name': 'EWB Password',      'Data Type': 'Password',       'required': False, 'Description / Notes': '',                                                    'Example Value': ''},
    {'Field Name': 'EWB API User ID',   'Data Type': 'Text',           'required': False, 'Description / Notes': 'For API-based E-Way Bill generation',                 'Example Value': 'DSNHP3424014000'},
    {'Field Name': 'EWB API Password',  'Data Type': 'Password',       'required': False, 'Description / Notes': '',                                                    'Example Value': 'Admin@2025'},
    {'_section': '── AUTHORISED SIGNATORY (for this registration)'},
    {'Field Name': 'Authorised Signatory','Data Type': 'Text',         'required': False, 'Description / Notes': 'Name of the authorised signatory for this GSTIN',     'Example Value': 'Rajesh Gupta'},
    {'Field Name': 'Signatory PAN',     'Data Type': 'Text',           'required': False, 'Description / Notes': 'PAN of the authorised signatory',                     'Example Value': 'ABCPG1234D'},
    {'_section': '── NOTES'},
    {'Field Name': 'Notes',             'Data Type': 'Long Text',      'required': False, 'Description / Notes': 'e.g. Filing frequency, special instructions',         'Example Value': ''},
]

# ─────────────────────────────────────────────
# SHEET 4 — DIRECTORS
# ─────────────────────────────────────────────
directors_data = [
    {'Field Name': 'Director ID',       'Data Type': 'Auto (UUID)',    'required': True,  'Description / Notes': 'System generated',                                    'Example Value': 'auto-generated'},
    {'Field Name': 'Client ID',         'Data Type': 'Link → Clients', 'required': True,  'Description / Notes': 'The company this person is a director of',            'Example Value': 'auto-linked'},
    {'Field Name': 'Person ID',         'Data Type': 'Link → Persons', 'required': True,  'Description / Notes': 'Links to the person record',                         'Example Value': 'auto-linked'},
    {'Field Name': 'Designation',       'Data Type': 'Dropdown',       'required': True,  'Description / Notes': 'Director | Managing Director | Whole-time Director | Independent Director | Nominee Director | Additional Director', 'Example Value': 'Director'},
    {'Field Name': 'DIN',               'Data Type': 'Text (8 char)',  'required': False, 'Description / Notes': 'Director Identification Number',                      'Example Value': '01234567'},
    {'Field Name': 'Date of Appointment','Data Type': 'Date',          'required': False, 'Description / Notes': '',                                                    'Example Value': '01/04/2015'},
    {'Field Name': 'Date of Cessation', 'Data Type': 'Date',           'required': False, 'Description / Notes': 'Leave blank if currently active',                    'Example Value': ''},
    {'Field Name': 'Is Active?',        'Data Type': 'Yes/No',         'required': True,  'Description / Notes': '',                                                    'Example Value': 'Yes'},
    {'Field Name': 'Is KMP?',           'Data Type': 'Yes/No',         'required': False, 'Description / Notes': 'Key Managerial Personnel',                            'Example Value': 'Yes'},
    {'Field Name': 'Notes',             'Data Type': 'Long Text',      'required': False, 'Description / Notes': '',                                                    'Example Value': ''},
]

# ─────────────────────────────────────────────
# SHEET 5 — SHAREHOLDERS
# ─────────────────────────────────────────────
shareholders_data = [
    {'Field Name': 'Shareholder ID',    'Data Type': 'Auto (UUID)',    'required': True,  'Description / Notes': 'System generated',                                    'Example Value': 'auto-generated'},
    {'Field Name': 'Client ID',         'Data Type': 'Link → Clients', 'required': True,  'Description / Notes': 'The company whose shares are held',                  'Example Value': 'auto-linked'},
    {'Field Name': 'Holder Type',       'Data Type': 'Dropdown',       'required': True,  'Description / Notes': 'Individual | Company | Trust | HUF | LLP',           'Example Value': 'Individual'},
    {'Field Name': 'Person ID',         'Data Type': 'Link → Persons', 'required': False, 'Description / Notes': 'Fill if Holder Type = Individual',                   'Example Value': 'auto-linked'},
    {'Field Name': 'Holding Entity ID', 'Data Type': 'Link → Clients', 'required': False, 'Description / Notes': 'Fill if Holder Type = Company/Trust/LLP',            'Example Value': ''},
    {'Field Name': 'Holder Name',       'Data Type': 'Text',           'required': True,  'Description / Notes': 'Name as per share records',                           'Example Value': 'Rajesh Kumar Gupta'},
    {'Field Name': 'PAN of Holder',     'Data Type': 'Text',           'required': False, 'Description / Notes': '',                                                    'Example Value': 'ABCPG1234D'},
    {'Field Name': 'Share Type',        'Data Type': 'Dropdown',       'required': False, 'Description / Notes': 'Equity | Preference | CCPS | OCPS',                  'Example Value': 'Equity'},
    {'Field Name': 'Number of Shares',  'Data Type': 'Number',         'required': False, 'Description / Notes': '',                                                    'Example Value': '5000'},
    {'Field Name': 'Face Value (₹)',    'Data Type': 'Number',         'required': False, 'Description / Notes': 'Per share face value',                                'Example Value': '10'},
    {'Field Name': 'Percentage %',      'Data Type': 'Decimal',        'required': False, 'Description / Notes': 'Shareholding percentage',                             'Example Value': '50.00'},
    {'Field Name': 'Date Acquired',     'Data Type': 'Date',           'required': False, 'Description / Notes': '',                                                    'Example Value': '01/04/2015'},
    {'Field Name': 'Is Active?',        'Data Type': 'Yes/No',         'required': True,  'Description / Notes': '',                                                    'Example Value': 'Yes'},
    {'Field Name': 'Notes',             'Data Type': 'Long Text',      'required': False, 'Description / Notes': '',                                                    'Example Value': ''},
]

# ─────────────────────────────────────────────
# SHEET 6 — PARTNERS (Firm / LLP)
# ─────────────────────────────────────────────
partners_data = [
    {'Field Name': 'Partner ID',        'Data Type': 'Auto (UUID)',    'required': True,  'Description / Notes': 'System generated',                                    'Example Value': 'auto-generated'},
    {'Field Name': 'Client ID',         'Data Type': 'Link → Clients', 'required': True,  'Description / Notes': 'The Firm or LLP',                                    'Example Value': 'auto-linked'},
    {'Field Name': 'Person ID',         'Data Type': 'Link → Persons', 'required': True,  'Description / Notes': 'The partner individual',                              'Example Value': 'auto-linked'},
    {'Field Name': 'Role',              'Data Type': 'Dropdown',       'required': True,  'Description / Notes': 'Partner | Designated Partner | Managing Partner | Sleeping Partner | Minor Partner', 'Example Value': 'Designated Partner'},
    {'Field Name': 'Profit Sharing Ratio %','Data Type': 'Decimal',   'required': False, 'Description / Notes': 'As per partnership deed / LLP agreement',             'Example Value': '50.00'},
    {'Field Name': 'Capital Contribution (₹)','Data Type': 'Number',  'required': False, 'Description / Notes': '',                                                    'Example Value': '500000'},
    {'Field Name': 'Date of Joining',   'Data Type': 'Date',           'required': False, 'Description / Notes': '',                                                    'Example Value': '01/04/2015'},
    {'Field Name': 'Date of Exit',      'Data Type': 'Date',           'required': False, 'Description / Notes': 'Leave blank if active',                               'Example Value': ''},
    {'Field Name': 'Is Active?',        'Data Type': 'Yes/No',         'required': True,  'Description / Notes': '',                                                    'Example Value': 'Yes'},
    {'Field Name': 'Notes',             'Data Type': 'Long Text',      'required': False, 'Description / Notes': '',                                                    'Example Value': ''},
]

# ─────────────────────────────────────────────
# SHEET 7 — BANK ACCOUNTS
# ─────────────────────────────────────────────
bank_data = [
    {'Field Name': 'Bank Account ID',   'Data Type': 'Auto (UUID)',    'required': True,  'Description / Notes': 'System generated',                                    'Example Value': 'auto-generated'},
    {'Field Name': 'Client ID',         'Data Type': 'Link → Clients', 'required': True,  'Description / Notes': 'The client this account belongs to',                 'Example Value': 'auto-linked'},
    {'Field Name': 'Bank Name',         'Data Type': 'Text',           'required': True,  'Description / Notes': '',                                                    'Example Value': 'State Bank of India'},
    {'Field Name': 'Account Number',    'Data Type': 'Text',           'required': True,  'Description / Notes': '',                                                    'Example Value': '20001791050000999'},
    {'Field Name': 'IFSC Code',         'Data Type': 'Text (11 char)', 'required': True,  'Description / Notes': 'Bank branch IFSC',                                    'Example Value': 'SBIN0001234'},
    {'Field Name': 'Branch Name',       'Data Type': 'Text',           'required': False, 'Description / Notes': '',                                                    'Example Value': 'Connaught Place'},
    {'Field Name': 'Account Type',      'Data Type': 'Dropdown',       'required': False, 'Description / Notes': 'Current | Savings | Cash Credit | Overdraft | EEFC', 'Example Value': 'Current'},
    {'Field Name': 'Is Primary?',       'Data Type': 'Yes/No',         'required': True,  'Description / Notes': 'Mark the main operating account',                    'Example Value': 'Yes'},
    {'Field Name': 'Net Banking User ID','Data Type': 'Text',          'required': False, 'Description / Notes': '',                                                    'Example Value': ''},
    {'Field Name': 'Net Banking Password','Data Type': 'Password',     'required': False, 'Description / Notes': '',                                                    'Example Value': ''},
    {'Field Name': 'Notes',             'Data Type': 'Long Text',      'required': False, 'Description / Notes': '',                                                    'Example Value': ''},
]

# ─────────────────────────────────────────────
# SHEET 8 — OTHER REGISTRATIONS
# ─────────────────────────────────────────────
other_reg_data = [
    {'Field Name': 'Registration ID',   'Data Type': 'Auto (UUID)',    'required': True,  'Description / Notes': 'System generated',                                    'Example Value': 'auto-generated'},
    {'Field Name': 'Client ID',         'Data Type': 'Link → Clients', 'required': True,  'Description / Notes': 'The client this registration belongs to',            'Example Value': 'auto-linked'},
    {'Field Name': 'Registration Type', 'Data Type': 'Dropdown',       'required': True,  'Description / Notes': 'TAN | EPF | ESI | MSME/Udyam | IEC | FSSAI | Professional Tax | Shops & Estab | Trade License | Drug License | Import Export Code | Others', 'Example Value': 'MSME/Udyam'},
    {'Field Name': 'Registration Number','Data Type': 'Text',          'required': True,  'Description / Notes': 'Unique registration/certificate number',              'Example Value': 'UDYAM-DL-01-0012345'},
    {'Field Name': 'Registration Date', 'Data Type': 'Date',           'required': False, 'Description / Notes': '',                                                    'Example Value': '15/06/2021'},
    {'Field Name': 'Valid Until',       'Data Type': 'Date',           'required': False, 'Description / Notes': 'For licenses with expiry — trigger alerts before expiry', 'Example Value': '14/06/2026'},
    {'Field Name': 'Issuing Authority', 'Data Type': 'Text',           'required': False, 'Description / Notes': 'Government dept that issued this',                    'Example Value': 'Ministry of MSME'},
    {'Field Name': 'State / Jurisdiction','Data Type': 'Dropdown',     'required': False, 'Description / Notes': 'Relevant for state-specific registrations',           'Example Value': 'Delhi'},
    {'Field Name': 'Portal User ID',    'Data Type': 'Text',           'required': False, 'Description / Notes': 'Login for this registration portal',                  'Example Value': ''},
    {'Field Name': 'Portal Password',   'Data Type': 'Password',       'required': False, 'Description / Notes': '',                                                    'Example Value': ''},
    {'Field Name': 'Is Active?',        'Data Type': 'Yes/No',         'required': True,  'Description / Notes': '',                                                    'Example Value': 'Yes'},
    {'Field Name': 'Notes',             'Data Type': 'Long Text',      'required': False, 'Description / Notes': '',                                                    'Example Value': ''},
]

# ─────────────────────────────────────────────
# BUILD ALL SHEETS
# ─────────────────────────────────────────────
wb.remove(wb.active)  # remove default sheet

sheets = [
    ("1 - Clients (Master)",      clients_data),
    ("2 - Persons (KYC)",         persons_data),
    ("3 - GST Registrations",     gst_data),
    ("4 - Directors",             directors_data),
    ("5 - Shareholders",          shareholders_data),
    ("6 - Partners (Firm-LLP)",   partners_data),
    ("7 - Bank Accounts",         bank_data),
    ("8 - Other Registrations",   other_reg_data),
]

for title, data in sheets:
    create_sheet(wb, title, COLS, data, WIDTHS)

# ─────────────────────────────────────────────
# LEGEND SHEET
# ─────────────────────────────────────────────
ws_legend = wb.create_sheet(title="README", index=0)
legend = [
    ("HOW TO USE THIS FILE", "1F3864", "FFFFFF", True, 14),
    ("", WHITE, "000000", False, 11),
    ("This file contains the proposed database schema for your CA firm's client management system.", WHITE, "000000", False, 11),
    ("Review each sheet, add/remove/rename fields, then share back.", WHITE, "000000", False, 11),
    ("", WHITE, "000000", False, 11),
    ("SHEET GUIDE", "1F3864", "FFFFFF", True, 12),
    ("Sheet 1 — Clients (Master)   : One row per company/individual/firm. Core identity, IT, MCA, EPF, ESI credentials.", "D6E4F0", "1F3864", False, 11),
    ("Sheet 2 — Persons (KYC)      : One row per individual person. Reused across Directors, Partners, Shareholders.", "D6E4F0", "1F3864", False, 11),
    ("Sheet 3 — GST Registrations  : One row per GSTIN. Multiple rows per client if they have multiple state registrations.", "D6E4F0", "1F3864", False, 11),
    ("Sheet 4 — Directors          : Links persons to companies with their designation and appointment dates.", "D6E4F0", "1F3864", False, 11),
    ("Sheet 5 — Shareholders       : Share ownership records. Holder can be an individual or another company.", "D6E4F0", "1F3864", False, 11),
    ("Sheet 6 — Partners (Firm/LLP): Links persons to firms/LLPs with profit sharing ratios.", "D6E4F0", "1F3864", False, 11),
    ("Sheet 7 — Bank Accounts      : Multiple bank accounts per client.", "D6E4F0", "1F3864", False, 11),
    ("Sheet 8 — Other Registrations: TAN, EPF, ESI, MSME, FSSAI, IEC, etc.", "D6E4F0", "1F3864", False, 11),
    ("", WHITE, "000000", False, 11),
    ("COLOUR CODING IN EACH SHEET", "1F3864", "FFFFFF", True, 12),
    ("Red field name   = Required (must fill)", WHITE, REQUIRED_COL, False, 11),
    ("Green field name = Optional (fill if available)", WHITE, OPTIONAL_COL, False, 11),
    ("", WHITE, "000000", False, 11),
    ("WHAT TO DO", "1F3864", "FFFFFF", True, 12),
    ("1. Add fields    — Add a new row with field name, type, and description.", WHITE, "000000", False, 11),
    ("2. Remove fields — Delete rows you don't need.", WHITE, "000000", False, 11),
    ("3. Rename fields — Change the Field Name to your preferred terminology.", WHITE, "000000", False, 11),
    ("4. Change types  — Update Data Type if needed.", WHITE, "000000", False, 11),
    ("5. Share back    — Send the modified file and we'll build the database accordingly.", WHITE, "000000", False, 11),
]

ws_legend.column_dimensions['A'].width = 90
for row_idx, (text, bg, fg, bold, size) in enumerate(legend, 1):
    cell = ws_legend.cell(row=row_idx, column=1, value=text)
    cell.fill = PatternFill("solid", fgColor=bg)
    cell.font = Font(bold=bold, color=fg, size=size)
    cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    ws_legend.row_dimensions[row_idx].height = 20

output_path = "/home/user/client-dashboard/CA_Client_Schema.xlsx"
wb.save(output_path)
print(f"Saved: {output_path}")
